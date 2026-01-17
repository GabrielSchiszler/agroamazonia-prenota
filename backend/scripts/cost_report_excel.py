#!/usr/bin/env python3
"""
Gera relatório Excel detalhado de custos por PDF processado
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Dados de custo
dados_com_textract = {
    'Textract': [
        {'Item': 'Páginas processadas', 'Quantidade': 1, 'Custo Unitário': 0.0015, 'Custo Total': 0.0015}
    ],
    'Bedrock': [
        {'Item': 'Chamadas parse_ocr', 'Quantidade': 1, 'Custo Unitário': 0.0015, 'Custo Total': 0.0015},
        {'Item': 'Chamadas validate_rules', 'Quantidade': 10, 'Custo Unitário': 0.0015, 'Custo Total': 0.0150}
    ],
    'Lambda': [
        {'Item': 'get_textract', 'Quantidade': 1, 'Custo Unitário': 0.000008, 'Custo Total': 0.000008},
        {'Item': 'processor', 'Quantidade': 1, 'Custo Unitário': 0.000021, 'Custo Total': 0.000021},
        {'Item': 'parse_ocr', 'Quantidade': 1, 'Custo Unitário': 0.000256, 'Custo Total': 0.000256},
        {'Item': 'validate_rules', 'Quantidade': 1, 'Custo Unitário': 0.000128, 'Custo Total': 0.000128},
        {'Item': 'notify_receipt', 'Quantidade': 1, 'Custo Unitário': 0.000003, 'Custo Total': 0.000003},
        {'Item': 'check_textract', 'Quantidade': 1, 'Custo Unitário': 0.000002, 'Custo Total': 0.000002},
        {'Item': 'send_to_protheus', 'Quantidade': 1, 'Custo Unitário': 0.000021, 'Custo Total': 0.000021},
        {'Item': 'update_metrics', 'Quantidade': 1, 'Custo Unitário': 0.000010, 'Custo Total': 0.000010},
        {'Item': 'Outras', 'Quantidade': 1, 'Custo Unitário': 0.000014, 'Custo Total': 0.000014}
    ],
    'DynamoDB': [
        {'Item': 'Read Units', 'Quantidade': 50, 'Custo Unitário': 0.000000005, 'Custo Total': 0.000013},
        {'Item': 'Write Units', 'Quantidade': 20, 'Custo Unitário': 0.000000025, 'Custo Total': 0.000001}
    ],
    'S3': [
        {'Item': 'Storage (30 dias)', 'Quantidade': 0.002, 'Custo Unitário': 0.000023, 'Custo Total': 0.000001},
        {'Item': 'PUT Requests', 'Quantidade': 1, 'Custo Unitário': 0.000005, 'Custo Total': 0.000003},
        {'Item': 'GET Requests', 'Quantidade': 2, 'Custo Unitário': 0.0000004, 'Custo Total': 0.000001}
    ],
    'Step Functions': [
        {'Item': 'Transições', 'Quantidade': 15, 'Custo Unitário': 0.000025, 'Custo Total': 0.000375}
    ]
}

dados_sem_textract = {
    'Textract': [
        {'Item': 'Páginas processadas', 'Quantidade': 0, 'Custo Unitário': 0.0015, 'Custo Total': 0.000000}
    ],
    'Bedrock': [
        {'Item': 'Chamadas parse_ocr', 'Quantidade': 0, 'Custo Unitário': 0.0015, 'Custo Total': 0.000000},
        {'Item': 'Chamadas validate_rules', 'Quantidade': 10, 'Custo Unitário': 0.0015, 'Custo Total': 0.015000}
    ],
    'Lambda': [
        {'Item': 'processor', 'Quantidade': 1, 'Custo Unitário': 0.000021, 'Custo Total': 0.000021},
        {'Item': 'parse_ocr', 'Quantidade': 1, 'Custo Unitário': 0.000256, 'Custo Total': 0.000256},
        {'Item': 'validate_rules', 'Quantidade': 1, 'Custo Unitário': 0.000128, 'Custo Total': 0.000128},
        {'Item': 'notify_receipt', 'Quantidade': 1, 'Custo Unitário': 0.000003, 'Custo Total': 0.000003},
        {'Item': 'check_textract', 'Quantidade': 1, 'Custo Unitário': 0.000002, 'Custo Total': 0.000002},
        {'Item': 'send_to_protheus', 'Quantidade': 1, 'Custo Unitário': 0.000021, 'Custo Total': 0.000021},
        {'Item': 'update_metrics', 'Quantidade': 1, 'Custo Unitário': 0.000010, 'Custo Total': 0.000010},
        {'Item': 'Outras', 'Quantidade': 1, 'Custo Unitário': 0.000014, 'Custo Total': 0.000014}
    ],
    'DynamoDB': [
        {'Item': 'Read Units', 'Quantidade': 50, 'Custo Unitário': 0.000000005, 'Custo Total': 0.000013},
        {'Item': 'Write Units', 'Quantidade': 20, 'Custo Unitário': 0.000000025, 'Custo Total': 0.000001}
    ],
    'S3': [
        {'Item': 'Storage (30 dias)', 'Quantidade': 0.002, 'Custo Unitário': 0.000023, 'Custo Total': 0.000001},
        {'Item': 'GET Requests', 'Quantidade': 2, 'Custo Unitário': 0.0000004, 'Custo Total': 0.000003}
    ],
    'Step Functions': [
        {'Item': 'Transições', 'Quantidade': 15, 'Custo Unitário': 0.000025, 'Custo Total': 0.000375}
    ]
}

def criar_planilha_detalhada(wb, nome_aba, dados, titulo):
    """Cria uma planilha detalhada com os custos"""
    ws = wb.create_sheet(title=nome_aba)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = titulo
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    linha_atual = 3
    
    # Cabeçalho geral
    ws[f'A{linha_atual}'] = 'Serviço AWS'
    ws[f'B{linha_atual}'] = 'Item'
    ws[f'C{linha_atual}'] = 'Quantidade'
    ws[f'D{linha_atual}'] = 'Custo Unitário ($)'
    ws[f'E{linha_atual}'] = 'Custo Total ($)'
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        cell = ws[f'{col}{linha_atual}']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    linha_atual += 1
    
    # Totais por serviço
    totais_servico = {}
    
    # Dados
    for servico, itens in dados.items():
        primeira_linha_servico = linha_atual
        
        for item in itens:
            ws[f'A{linha_atual}'] = servico
            ws[f'B{linha_atual}'] = item['Item']
            ws[f'C{linha_atual}'] = item['Quantidade']
            ws[f'D{linha_atual}'] = item['Custo Unitário']
            ws[f'E{linha_atual}'] = item['Custo Total']
            
            # Formatação
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}{linha_atual}']
                cell.border = border
                if col in ['D', 'E']:
                    cell.number_format = '$#,##0.000000'
                elif col == 'C':
                    cell.number_format = '#,##0.0000'
            
            # Acumular total do serviço
            if servico not in totais_servico:
                totais_servico[servico] = 0
            totais_servico[servico] += item['Custo Total']
            
            linha_atual += 1
        
        # Linha de subtotal do serviço
        ws[f'A{linha_atual}'] = f'SUBTOTAL {servico}'
        ws[f'E{linha_atual}'] = totais_servico[servico]
        ws[f'A{linha_atual}'].font = Font(bold=True)
        ws[f'E{linha_atual}'].font = Font(bold=True)
        ws[f'E{linha_atual}'].number_format = '$#,##0.000000'
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{linha_atual}']
            cell.border = border
            if col == 'A':
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        linha_atual += 1
    
    # Linha de total geral
    total_geral = sum(totais_servico.values())
    ws[f'A{linha_atual}'] = 'TOTAL GERAL'
    ws[f'E{linha_atual}'] = total_geral
    ws[f'A{linha_atual}'].font = Font(bold=True, size=12)
    ws[f'E{linha_atual}'].font = Font(bold=True, size=12)
    ws[f'E{linha_atual}'].number_format = '$#,##0.000000'
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        cell = ws[f'{col}{linha_atual}']
        cell.border = border
        if col == 'A':
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
        if col == 'E':
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

def criar_planilha_resumo(wb):
    """Cria planilha de resumo comparativo"""
    ws = wb.create_sheet(title='Resumo Comparativo', index=0)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = 'RESUMO COMPARATIVO - CUSTO POR PDF PROCESSADO'
    ws['A1'].font = title_font
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    linha = 3
    
    # Cabeçalho
    ws[f'A{linha}'] = 'Serviço AWS'
    ws[f'B{linha}'] = 'Com Textract ($)'
    ws[f'C{linha}'] = 'Sem Textract ($)'
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{linha}']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    linha += 1
    
    # Calcular totais
    totais_com = {}
    totais_sem = {}
    
    for servico in dados_com_textract.keys():
        totais_com[servico] = sum(item['Custo Total'] for item in dados_com_textract[servico])
        totais_sem[servico] = sum(item['Custo Total'] for item in dados_sem_textract[servico])
    
    # Dados
    for servico in totais_com.keys():
        ws[f'A{linha}'] = servico
        ws[f'B{linha}'] = totais_com[servico]
        ws[f'C{linha}'] = totais_sem[servico]
        
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{linha}']
            cell.border = border
            if col in ['B', 'C']:
                cell.number_format = '$#,##0.000000'
        
        linha += 1
    
    # Linha de diferença
    ws[f'A{linha}'] = 'DIFERENÇA'
    ws[f'B{linha}'] = sum(totais_com.values()) - sum(totais_sem.values())
    ws[f'C{linha}'] = '-'
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{linha}']
        cell.border = border
        cell.font = Font(bold=True)
        if col == 'B':
            cell.number_format = '$#,##0.000000'
    
    linha += 1
    
    # Total
    total_com = sum(totais_com.values())
    total_sem = sum(totais_sem.values())
    
    ws[f'A{linha}'] = 'TOTAL'
    ws[f'B{linha}'] = total_com
    ws[f'C{linha}'] = total_sem
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{linha}']
        cell.border = border
        cell.font = Font(bold=True, size=12)
        if col in ['B', 'C']:
            cell.number_format = '$#,##0.000000'
        if col == 'A':
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
        if col in ['B', 'C']:
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
    
    # Ajustar largura
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Adicionar projeção mensal
    linha += 3
    ws[f'A{linha}'] = 'PROJEÇÃO MENSAL (1.000 PDFs)'
    ws[f'A{linha}'].font = title_font
    ws.merge_cells(f'A{linha}:C{linha}')
    
    linha += 1
    ws[f'A{linha}'] = 'Cenário'
    ws[f'B{linha}'] = 'Custo por PDF'
    ws[f'C{linha}'] = 'Custo Mensal'
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{linha}']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    linha += 1
    ws[f'A{linha}'] = '100% com Textract'
    ws[f'B{linha}'] = total_com
    ws[f'C{linha}'] = total_com * 1000
    linha += 1
    ws[f'A{linha}'] = '100% sem Textract'
    ws[f'B{linha}'] = total_sem
    ws[f'C{linha}'] = total_sem * 1000
    linha += 1
    ws[f'A{linha}'] = 'Mix 70/30'
    ws[f'B{linha}'] = (total_com * 0.7 + total_sem * 0.3)
    ws[f'C{linha}'] = (total_com * 0.7 + total_sem * 0.3) * 1000
    
    for row in range(linha - 2, linha + 1):
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.border = border
            if col in ['B', 'C']:
                cell.number_format = '$#,##0.00'

def main():
    """Função principal"""
    wb = Workbook()
    
    # Remover sheet padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Criar planilhas
    criar_planilha_resumo(wb)
    criar_planilha_detalhada(wb, 'Com Textract', dados_com_textract, 
                            'CUSTO DETALHADO - PDF COM TEXTRACT')
    criar_planilha_detalhada(wb, 'Sem Textract', dados_sem_textract, 
                            'CUSTO DETALHADO - PDF SEM TEXTRACT')
    
    # Salvar arquivo
    nome_arquivo = 'custo_por_pdf_agroamazonia.xlsx'
    wb.save(nome_arquivo)
    print(f"✅ Arquivo Excel gerado: {nome_arquivo}")
    print(f"   - Resumo Comparativo")
    print(f"   - Com Textract (detalhado)")
    print(f"   - Sem Textract (detalhado)")

if __name__ == '__main__':
    main()